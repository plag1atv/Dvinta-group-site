import json
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent.parent

EXCEL_PATH = BASE_DIR / "Прейскурант ДВИНТА 2026.xlsx"
OUTPUT_DIR = BASE_DIR / "app" / "static" / "data"
OUTPUT_PATH = OUTPUT_DIR / "price_items.json"


SECTION_NAMES = {
    "27": "27. Геометрические величины",
    "28": "28. Механические величины",
    "29": "29. Параметры потока, расхода, уровня, объема веществ",
    "30": "30. Давление и вакуумные измерения",
    "31": "31. Физико-химический состав и свойства веществ",
    "32": "32. Теплофизические и температурные измерения",
    "33": "33. Измерения времени и частоты",
    "44": "44. Элементы измерительных систем",
}


def clean_value(value):
    if value is None:
        return ""

    return str(value).replace("\t", " ").strip()


def make_item(
    item_id,
    code,
    section,
    name,
    range_value,
    verification_price,
    calibration_price,
    note,
):
    return {
        "id": item_id,
        "code": clean_value(code),
        "section": section,
        "name": clean_value(name),
        "range": clean_value(range_value),
        "verification_price": clean_value(verification_price),
        "calibration_price": clean_value(calibration_price),
        "note": clean_value(note),
    }


def parse_price_sheet(sheet, section, start_id):
    items = []
    current_name = ""
    item_id = start_id

    for row in sheet.iter_rows(min_row=6, values_only=True):
        code = clean_value(row[0])
        name = clean_value(row[1])
        range_value = clean_value(row[2])
        verification_price = clean_value(row[3])
        calibration_price = clean_value(row[4])
        note = clean_value(row[5]) if len(row) > 5 else ""

        if not code:
            continue

        if name:
            current_name = name

        if not current_name:
            continue

        item = make_item(
            item_id=item_id,
            code=code,
            section=section,
            name=current_name,
            range_value=range_value,
            verification_price=verification_price,
            calibration_price=calibration_price,
            note=note,
        )

        items.append(item)
        item_id += 1

    return items, item_id


def parse_extra_charges_sheet(sheet, start_id):
    items = []
    item_id = start_id

    for row in sheet.iter_rows(min_row=6, values_only=True):
        code = clean_value(row[0])
        name = clean_value(row[1])
        price = clean_value(row[2])

        if not code or not name:
            continue

        item = {
            "id": item_id,
            "code": code,
            "section": "Доплаты",
            "name": name,
            "range": "",
            "verification_price": price,
            "calibration_price": "",
            "note": "Доплата применяется к основной стоимости работ",
        }

        items.append(item)
        item_id += 1

    return items, item_id


def main():
    if not EXCEL_PATH.exists():
        print(f"Файл не найден: {EXCEL_PATH}")
        return

    workbook = load_workbook(EXCEL_PATH, data_only=True)

    price_items = []
    item_id = 1

    for sheet_name, section_name in SECTION_NAMES.items():
        if sheet_name not in workbook.sheetnames:
            print(f"Лист не найден: {sheet_name}")
            continue

        sheet_items, item_id = parse_price_sheet(
            sheet=workbook[sheet_name],
            section=section_name,
            start_id=item_id,
        )

        price_items.extend(sheet_items)

    if "Доплаты" in workbook.sheetnames:
        extra_items, item_id = parse_extra_charges_sheet(
            sheet=workbook["Доплаты"],
            start_id=item_id,
        )

        price_items.extend(extra_items)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    with OUTPUT_PATH.open("w", encoding="utf-8") as file:
        json.dump(price_items, file, ensure_ascii=False, indent=2)

    print(f"Готово. Создан файл: {OUTPUT_PATH}")
    print(f"Количество позиций: {len(price_items)}")


if __name__ == "__main__":
    main()